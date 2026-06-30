"""Importador de Excel de productos y precios de proveedor.

Lee el archivo Excel consolidado de productos/comparativa de Index Clima
y lo importa a la base de datos manteniendo historial completo.

Uso CLI:
    python -m quote_engine.importers.excel_products_importer --file data/raw/archivo.xlsx

Uso programático:
    from quote_engine.importers.excel_products_importer import ExcelProductsImporter
    result = ExcelProductsImporter().run("data/raw/archivo.xlsx")
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Normalización de texto
# ---------------------------------------------------------------------------

def normalize_text(text: str) -> str:
    """Normaliza texto para búsqueda: minúsculas, sin tildes, sin caracteres especiales."""
    if not text:
        return ""
    text = str(text).strip()
    # Eliminar tildes
    text = "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )
    text = text.lower()
    # Colapsar espacios
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_supplier_name(name: str) -> str:
    """Normaliza nombre de proveedor."""
    normalized = normalize_text(name)
    # Eliminar sufijos comunes
    for suffix in [" s.l.", " s.a.", " s.l.u.", " sl", " sa"]:
        if normalized.endswith(suffix):
            normalized = normalized[:-len(suffix)].strip()
    return normalized


# ---------------------------------------------------------------------------
# Estructuras de resultado
# ---------------------------------------------------------------------------

@dataclass
class ImportedSupplier:
    name: str
    normalized_name: str
    is_new: bool = False
    db_id: Optional[int] = None


@dataclass
class ImportedProduct:
    description: str
    normalized_description: str
    category: Optional[str] = None
    internal_code: Optional[str] = None
    supplier_reference: Optional[str] = None
    unit: Optional[str] = None
    is_new: bool = False
    db_id: Optional[int] = None


@dataclass
class ImportedPrice:
    product_description: str
    supplier_name: str
    net_unit_price: float
    gross_unit_price: Optional[float] = None
    discount_percent: Optional[float] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    document_number: Optional[str] = None
    document_date: Optional[date] = None
    source_sheet: Optional[str] = None
    source_row: Optional[int] = None
    confidence: str = "media"
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportReport:
    source_file: str
    source_hash: str
    imported_at: datetime = field(default_factory=datetime.now)

    total_rows: int = 0
    created_products: int = 0
    updated_products: int = 0
    created_prices: int = 0
    warnings_count: int = 0
    errors_count: int = 0

    suppliers_detected: list[str] = field(default_factory=list)
    products_detected: list[str] = field(default_factory=list)
    prices_created: list[dict] = field(default_factory=list)
    prices_updated: list[dict] = field(default_factory=list)
    dubious_products: list[dict] = field(default_factory=list)
    ignored_rows: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    no_code_products: list[str] = field(default_factory=list)
    dubious_units: list[dict] = field(default_factory=list)
    negative_prices: list[dict] = field(default_factory=list)
    possible_credits: list[dict] = field(default_factory=list)
    possible_duplicates: list[dict] = field(default_factory=list)

    def add_warning(self, msg: str) -> None:
        self.warnings.append(msg)
        self.warnings_count += 1
        logger.warning(msg)

    def add_error(self, msg: str) -> None:
        self.errors.append(msg)
        self.errors_count += 1
        logger.error(msg)

    def to_markdown(self) -> str:
        lines = [
            f"# Informe de importación — {self.source_file}",
            f"> Importado: {self.imported_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"> Hash: `{self.source_hash[:16]}...`",
            "",
            "## Resumen",
            "",
            f"- Filas procesadas: **{self.total_rows}**",
            f"- Productos nuevos: **{self.created_products}**",
            f"- Productos actualizados: **{self.updated_products}**",
            f"- Precios creados: **{self.created_prices}**",
            f"- Warnings: **{self.warnings_count}**",
            f"- Errores: **{self.errors_count}**",
            "",
        ]

        if self.suppliers_detected:
            lines += ["## Proveedores detectados", ""]
            for s in self.suppliers_detected:
                lines.append(f"- {s}")
            lines.append("")

        if self.products_detected:
            lines += [f"## Productos detectados ({len(self.products_detected)})", ""]
            for p in self.products_detected[:50]:
                lines.append(f"- {p}")
            if len(self.products_detected) > 50:
                lines.append(f"- ... y {len(self.products_detected) - 50} más")
            lines.append("")

        if self.dubious_products:
            lines += ["## Productos dudosos", ""]
            for d in self.dubious_products:
                lines.append(f"- `{d.get('description', '?')}` — {d.get('reason', '?')}")
            lines.append("")

        if self.no_code_products:
            lines += ["## Productos sin código interno", ""]
            for p in self.no_code_products:
                lines.append(f"- {p}")
            lines.append("")

        if self.dubious_units:
            lines += ["## Unidades dudosas", ""]
            for d in self.dubious_units:
                lines.append(f"- `{d.get('product', '?')}` — unidad: `{d.get('unit', '?')}` — {d.get('reason', '?')}")
            lines.append("")

        if self.negative_prices:
            lines += ["## Precios negativos (posibles abonos)", ""]
            for d in self.negative_prices:
                lines.append(f"- `{d.get('product', '?')}` — precio: {d.get('price', '?')}")
            lines.append("")

        if self.possible_duplicates:
            lines += ["## Posibles duplicados", ""]
            for d in self.possible_duplicates:
                lines.append(f"- `{d.get('description', '?')}` — {d.get('reason', '?')}")
            lines.append("")

        if self.warnings:
            lines += ["## Warnings", ""]
            for w in self.warnings:
                lines.append(f"- ⚠ {w}")
            lines.append("")

        if self.errors:
            lines += ["## Errores", ""]
            for e in self.errors:
                lines.append(f"- ❌ {e}")
            lines.append("")

        if self.ignored_rows:
            lines += [f"## Filas ignoradas ({len(self.ignored_rows)})", ""]
            for r in self.ignored_rows[:20]:
                lines.append(f"- Fila {r.get('row', '?')}: {r.get('reason', '?')}")
            lines.append("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Detectores de columnas
# ---------------------------------------------------------------------------

# Columnas que indican precio neto
_PRICE_NET_KEYS = {
    "precio neto", "precio net", "p. neto", "p.neto", "net", "neto",
    "precio unitario neto", "coste neto", "cost", "precio final",
}
# Columnas que indican precio bruto/tarifa
_PRICE_GROSS_KEYS = {
    "tarifa", "pvp", "precio bruto", "precio tarifa", "bruto", "precio pvp",
}
# Columnas que indican descuento
_DISCOUNT_KEYS = {
    "dto", "descuento", "dto%", "desc%", "discount", "% dto", "descuento %",
}
# Columnas de descripción de producto
_DESC_KEYS = {
    "descripcion", "descripción", "articulo", "artículo", "producto",
    "referencia", "concepto", "item", "material", "nombre",
}
# Columnas de unidad
_UNIT_KEYS = {"unidad", "ud", "u.", "unid.", "unidades", "unit"}
# Columnas de cantidad
_QTY_KEYS = {"cantidad", "qty", "q", "cant.", "uds", "unidades"}
# Columnas de fecha
_DATE_KEYS = {"fecha", "date", "fecha factura", "fecha doc", "f.factura"}
# Columnas de número de documento
_DOC_KEYS = {"factura", "albarán", "albaran", "nº factura", "numero", "doc", "ref"}
# Columnas de proveedor
_SUPPLIER_KEYS = {"proveedor", "supplier", "suministrador", "fabricante"}


def _match_col(header: str, keywords: set[str]) -> bool:
    """True si el header normalizado contiene alguna keyword."""
    h = normalize_text(str(header))
    return any(kw in h for kw in keywords)


def _detect_columns(headers: list[Any]) -> dict[str, int]:
    """Detecta índices de columnas por tipo."""
    result: dict[str, list[int]] = {
        "description": [], "price_net": [], "price_gross": [],
        "discount": [], "unit": [], "quantity": [], "date": [],
        "document": [], "supplier": [],
    }

    for i, h in enumerate(headers):
        if not h or str(h).strip() == "":
            continue
        h_str = str(h).strip()
        if _match_col(h_str, _DESC_KEYS):
            result["description"].append(i)
        if _match_col(h_str, _PRICE_NET_KEYS):
            result["price_net"].append(i)
        if _match_col(h_str, _PRICE_GROSS_KEYS):
            result["price_gross"].append(i)
        if _match_col(h_str, _DISCOUNT_KEYS):
            result["discount"].append(i)
        if _match_col(h_str, _UNIT_KEYS):
            result["unit"].append(i)
        if _match_col(h_str, _QTY_KEYS):
            result["quantity"].append(i)
        if _match_col(h_str, _DATE_KEYS):
            result["date"].append(i)
        if _match_col(h_str, _DOC_KEYS):
            result["document"].append(i)
        if _match_col(h_str, _SUPPLIER_KEYS):
            result["supplier"].append(i)

    # Tomar el primero de cada tipo
    return {k: v[0] if v else -1 for k, v in result.items()}


# ---------------------------------------------------------------------------
# Lectura de valores
# ---------------------------------------------------------------------------

def _safe_float(value: Any) -> Optional[float]:
    """Convierte valor de celda a float. None si no es número."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".").replace("€", "").replace("%", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _safe_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _safe_date(value: Any) -> Optional[date]:
    """Convierte valor de celda a date."""
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Importador principal
# ---------------------------------------------------------------------------

class ExcelProductsImporter:
    """Importa productos y precios desde un archivo Excel."""

    # Cambio de precio que dispara advertencia (%)
    PRICE_CHANGE_THRESHOLD = 0.30  # 30%

    def __init__(self, db_session=None):
        """
        db_session: sesión SQLAlchemy opcional. Si es None, opera en modo dry-run
                    (genera informe pero no escribe en DB).
        """
        self.session = db_session
        self._suppliers_cache: dict[str, Any] = {}  # normalized_name -> ORM object or dict
        self._products_cache: dict[str, Any] = {}   # normalized_description -> ORM object or dict

    def _compute_hash(self, path: Path) -> str:
        """Calcula SHA-256 del archivo."""
        sha256 = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                sha256.update(chunk)
        return sha256.hexdigest()

    def _is_already_imported(self, file_hash: str) -> bool:
        """Comprueba si este archivo ya fue importado (por hash)."""
        if self.session is None:
            return False
        try:
            from quote_engine.db.models import PriceImportBatch
            existing = self.session.query(PriceImportBatch).filter_by(
                source_hash=file_hash
            ).first()
            return existing is not None
        except Exception:
            return False

    def _get_or_create_supplier(self, name: str, report: ImportReport) -> Optional[Any]:
        """Obtiene o crea un proveedor."""
        normalized = normalize_supplier_name(name)
        if not normalized:
            return None

        if normalized in self._suppliers_cache:
            return self._suppliers_cache[normalized]

        if self.session is not None:
            try:
                from quote_engine.db.models import Supplier
                supplier = self.session.query(Supplier).filter_by(
                    normalized_name=normalized
                ).first()
                if supplier is None:
                    supplier = Supplier(name=name, normalized_name=normalized)
                    self.session.add(supplier)
                    self.session.flush()
                    report.add_warning(f"Nuevo proveedor creado: {name!r}")
                self._suppliers_cache[normalized] = supplier
                return supplier
            except Exception as exc:
                report.add_error(f"Error al crear proveedor {name!r}: {exc}")
                return None
        else:
            # Modo dry-run
            obj = {"name": name, "normalized_name": normalized, "id": None}
            self._suppliers_cache[normalized] = obj
            return obj

    def _get_or_create_product(
        self,
        description: str,
        category: Optional[str],
        unit: Optional[str],
        internal_code: Optional[str],
        supplier_reference: Optional[str],
        report: ImportReport,
    ) -> Optional[Any]:
        """Obtiene o crea un producto."""
        normalized = normalize_text(description)
        if not normalized:
            return None

        # Detectar duplicados potenciales
        if normalized in self._products_cache:
            cached = self._products_cache[normalized]
            # Si cambia la unidad, advertir
            cached_unit = cached.get("unit_calc") if isinstance(cached, dict) else getattr(cached, "unit_calc", None)
            if unit and cached_unit and normalize_text(unit) != normalize_text(str(cached_unit)):
                report.dubious_units.append({
                    "product": description,
                    "unit": unit,
                    "reason": f"Unidad actual: {unit!r}, unidad en cache: {cached_unit!r}",
                })
            return cached

        if self.session is not None:
            try:
                from quote_engine.db.models import Product
                product = self.session.query(Product).filter_by(
                    normalized_description=normalized
                ).first()
                if product is None:
                    product = Product(
                        description=description,
                        normalized_description=normalized,
                        category=category,
                        unit_calc=unit,
                        internal_code=internal_code,
                        supplier_reference=supplier_reference,
                    )
                    self.session.add(product)
                    self.session.flush()
                    report.created_products += 1
                    if not internal_code:
                        report.no_code_products.append(description)
                else:
                    report.updated_products += 1
                self._products_cache[normalized] = product
                return product
            except Exception as exc:
                report.add_error(f"Error al crear producto {description!r}: {exc}")
                return None
        else:
            # Modo dry-run
            obj = {
                "description": description,
                "normalized_description": normalized,
                "category": category,
                "unit_calc": unit,
                "id": None,
            }
            if not internal_code:
                report.no_code_products.append(description)
            self._products_cache[normalized] = obj
            return obj

    def _check_price_anomaly(
        self,
        product_id: Optional[int],
        new_price: float,
        report: ImportReport,
        description: str,
    ) -> None:
        """Detecta cambios de precio anómalos respecto al histórico."""
        if self.session is None or product_id is None:
            return
        try:
            from quote_engine.db.models import SupplierPrice
            last = (
                self.session.query(SupplierPrice)
                .filter_by(product_id=product_id, is_current=True)
                .order_by(SupplierPrice.created_at.desc())
                .first()
            )
            if last and last.net_unit_price > 0:
                change = abs(new_price - last.net_unit_price) / last.net_unit_price
                if change > self.PRICE_CHANGE_THRESHOLD:
                    report.add_warning(
                        f"Cambio de precio >30% en {description!r}: "
                        f"{last.net_unit_price:.2f} → {new_price:.2f} "
                        f"({change * 100:.1f}%)"
                    )
        except Exception:
            pass

    def _save_price(
        self,
        product: Any,
        supplier: Any,
        price_data: ImportedPrice,
        source_file: str,
        report: ImportReport,
    ) -> None:
        """Guarda un precio en la base de datos."""
        if self.session is None:
            report.prices_created.append({
                "product": price_data.product_description,
                "supplier": price_data.supplier_name,
                "price": price_data.net_unit_price,
                "confidence": price_data.confidence,
            })
            report.created_prices += 1
            return

        product_id = product.id if hasattr(product, "id") else product.get("id")
        supplier_id = supplier.id if hasattr(supplier, "id") else supplier.get("id")

        if product_id is None or supplier_id is None:
            report.add_warning(
                f"No se pudo guardar precio para {price_data.product_description!r}: IDs no disponibles"
            )
            return

        # Comprobar anomalía de precio
        self._check_price_anomaly(product_id, price_data.net_unit_price, report, price_data.product_description)

        try:
            from quote_engine.db.models import SupplierPrice

            # Marcar precios anteriores de este producto+proveedor como no actuales
            self.session.query(SupplierPrice).filter_by(
                product_id=product_id,
                supplier_id=supplier_id,
                is_current=True,
            ).update({"is_current": False})

            # Crear nuevo precio
            price_obj = SupplierPrice(
                product_id=product_id,
                supplier_id=supplier_id,
                source_type="excel",
                source_file=source_file,
                source_sheet=price_data.source_sheet,
                source_row=price_data.source_row,
                document_number=price_data.document_number,
                document_date=datetime.combine(price_data.document_date, datetime.min.time())
                    if price_data.document_date else None,
                quantity=price_data.quantity,
                gross_unit_price=price_data.gross_unit_price,
                discount_percent=price_data.discount_percent,
                net_unit_price=price_data.net_unit_price,
                confidence=price_data.confidence,
                is_current=True,
            )
            self.session.add(price_obj)
            self.session.flush()
            report.created_prices += 1
        except Exception as exc:
            report.add_error(
                f"Error al guardar precio para {price_data.product_description!r}: {exc}"
            )

    def _process_sheet(
        self,
        ws,
        sheet_name: str,
        supplier_name: Optional[str],
        report: ImportReport,
    ) -> list[ImportedPrice]:
        """Procesa una hoja del Excel y devuelve lista de precios."""
        prices: list[ImportedPrice] = []

        # Obtener filas como listas
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            report.add_warning(f"Hoja {sheet_name!r} vacía")
            return prices

        # Buscar fila de cabecera (primera fila no vacía)
        header_row_idx = 0
        for i, row in enumerate(rows):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row_idx = i
                break

        headers = rows[header_row_idx]
        cols = _detect_columns(headers)

        # Si no hay columna de descripción ni de precio, saltar hoja
        if cols["description"] == -1 and cols["price_net"] == -1:
            report.ignored_rows.append({
                "row": 0,
                "reason": f"Hoja {sheet_name!r}: no se detectaron columnas de descripción o precio",
            })
            return prices

        # Detectar proveedor desde cabecera de hoja si no viene externo
        if supplier_name is None and cols["supplier"] != -1:
            # Intentar extraer de primera fila de datos
            for row in rows[header_row_idx + 1:header_row_idx + 3]:
                val = _safe_str(row[cols["supplier"]])
                if val:
                    supplier_name = val
                    break

        # Si aún sin proveedor, usar nombre de hoja
        if supplier_name is None:
            supplier_name = sheet_name

        # Procesar filas de datos
        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            report.total_rows += 1

            # Descripción
            desc = _safe_str(row[cols["description"]]) if cols["description"] != -1 else None
            if not desc:
                report.ignored_rows.append({"row": row_idx, "reason": "Sin descripción"})
                continue

            # Precio neto
            price_net = _safe_float(row[cols["price_net"]]) if cols["price_net"] != -1 else None
            price_gross = _safe_float(row[cols["price_gross"]]) if cols["price_gross"] != -1 else None
            discount = _safe_float(row[cols["discount"]]) if cols["discount"] != -1 else None

            # Calcular precio neto si tenemos bruto y descuento
            if price_net is None and price_gross is not None and discount is not None:
                price_net = price_gross * (1 - discount / 100.0)

            if price_net is None:
                report.ignored_rows.append({"row": row_idx, "reason": f"Sin precio neto en {desc!r}"})
                continue

            # Precios negativos
            if price_net < 0:
                report.negative_prices.append({"product": desc, "price": price_net, "row": row_idx})
                report.possible_credits.append({"product": desc, "price": price_net, "row": row_idx})
                report.add_warning(f"Precio negativo en fila {row_idx}: {desc!r} = {price_net}")
                continue

            # Unidad
            unit = _safe_str(row[cols["unit"]]) if cols["unit"] != -1 else None
            # Detectar unidades dudosas
            if unit:
                unit_norm = normalize_text(unit)
                if unit_norm in {"rollo", "bobina", "caja", "pack"}:
                    report.dubious_units.append({
                        "product": desc,
                        "unit": unit,
                        "reason": "Unidad de empaque — puede necesitar conversion_factor",
                    })

            # Cantidad
            qty = _safe_float(row[cols["quantity"]]) if cols["quantity"] != -1 else None

            # Fecha y documento
            doc_date = _safe_date(row[cols["date"]]) if cols["date"] != -1 else None
            doc_num = _safe_str(row[cols["document"]]) if cols["document"] != -1 else None

            # Determinar confianza
            confidence = "alta"
            if price_gross is None and discount is None:
                confidence = "media"  # Solo precio neto, sin trazabilidad de descuento
            if doc_date is None:
                confidence = "media"

            imported_price = ImportedPrice(
                product_description=desc,
                supplier_name=supplier_name,
                net_unit_price=round(price_net, 6),
                gross_unit_price=price_gross,
                discount_percent=discount,
                quantity=qty,
                unit=unit,
                document_number=doc_num,
                document_date=doc_date,
                source_sheet=sheet_name,
                source_row=row_idx,
                confidence=confidence,
            )
            prices.append(imported_price)

        return prices

    def run(
        self,
        file_path: str | Path,
        force: bool = False,
        report_output_dir: Optional[str] = None,
    ) -> ImportReport:
        """Ejecuta la importación completa.

        Args:
            file_path: Ruta al archivo Excel.
            force: Si True, importa aunque el hash ya esté en la DB.
            report_output_dir: Directorio donde guardar el informe MD.

        Returns:
            ImportReport con el resultado completo.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl no está instalado. Ejecuta: pip install openpyxl"
            )

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {path}")

        file_hash = self._compute_hash(path)
        report = ImportReport(source_file=str(path), source_hash=file_hash)

        # Comprobar si ya importado
        if not force and self._is_already_imported(file_hash):
            report.add_warning(
                f"Este archivo ya fue importado (hash: {file_hash[:16]}...). "
                "Usa --force para reimportar."
            )
            return report

        # Cargar workbook
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            report.add_error(f"No se pudo leer el Excel: {exc}")
            return report

        sheet_names = wb.sheetnames
        logger.info(f"Hojas encontradas: {sheet_names}")

        all_prices: list[ImportedPrice] = []

        # Hojas a ignorar
        IGNORE_SHEETS = {"indice", "índice", "portada", "cover", "summary", "resumen"}

        for sheet_name in sheet_names:
            if normalize_text(sheet_name) in IGNORE_SHEETS:
                continue

            ws = wb[sheet_name]

            # Intentar detectar proveedor desde el nombre de la hoja
            supplier_from_sheet: Optional[str] = None
            sheet_norm = normalize_text(sheet_name)
            if any(kw in sheet_norm for kw in ["frigicoll", "airwell", "daitsu", "haier", "mitsubishi", "carrier", "daikin", "toshiba"]):
                supplier_from_sheet = sheet_name

            prices = self._process_sheet(ws, sheet_name, supplier_from_sheet, report)
            all_prices.extend(prices)

        wb.close()

        # Detectar proveedores y productos únicos
        suppliers_set = sorted({p.supplier_name for p in all_prices})
        products_set = sorted({p.product_description for p in all_prices})

        report.suppliers_detected = suppliers_set
        report.products_detected = products_set

        # Detectar posibles duplicados de producto
        normalized_descs: dict[str, list[str]] = {}
        for desc in products_set:
            nd = normalize_text(desc)
            normalized_descs.setdefault(nd, []).append(desc)
        for nd, descs in normalized_descs.items():
            if len(descs) > 1:
                report.possible_duplicates.append({
                    "description": nd,
                    "variants": descs,
                    "reason": f"{len(descs)} descripciones similares",
                })

        # Guardar en DB (o dry-run)
        if self.session is not None:
            try:
                for price_data in all_prices:
                    supplier = self._get_or_create_supplier(price_data.supplier_name, report)
                    if supplier is None:
                        continue

                    product = self._get_or_create_product(
                        description=price_data.product_description,
                        category=None,
                        unit=price_data.unit,
                        internal_code=None,
                        supplier_reference=None,
                        report=report,
                    )
                    if product is None:
                        continue

                    self._save_price(product, supplier, price_data, str(path), report)

                # Guardar batch
                from quote_engine.db.models import PriceImportBatch
                batch = PriceImportBatch(
                    source_file=str(path),
                    source_hash=file_hash,
                    status="completed",
                    total_rows=report.total_rows,
                    created_products=report.created_products,
                    updated_products=report.updated_products,
                    created_prices=report.created_prices,
                    warnings_count=report.warnings_count,
                    errors_count=report.errors_count,
                )
                self.session.add(batch)
                self.session.commit()

            except Exception as exc:
                report.add_error(f"Error durante importación DB: {exc}")
                if self.session:
                    self.session.rollback()
        else:
            # Dry-run: solo registrar estadísticas
            for price_data in all_prices:
                supplier_key = normalize_supplier_name(price_data.supplier_name)
                if supplier_key not in self._suppliers_cache:
                    self._suppliers_cache[supplier_key] = {"name": price_data.supplier_name}

                product_key = normalize_text(price_data.product_description)
                if product_key not in self._products_cache:
                    self._products_cache[product_key] = {"description": price_data.product_description}
                    report.created_products += 1

                report.created_prices += 1
                report.prices_created.append({
                    "product": price_data.product_description,
                    "supplier": price_data.supplier_name,
                    "price": price_data.net_unit_price,
                    "confidence": price_data.confidence,
                })

        # Guardar informe
        report_dir = Path(report_output_dir) if report_output_dir else Path("data/reports")
        report_dir.mkdir(parents=True, exist_ok=True)
        timestamp = report.imported_at.strftime("%Y%m%d_%H%M%S")
        report_path = report_dir / f"import_report_{timestamp}.md"
        report_path.write_text(report.to_markdown(), encoding="utf-8")
        logger.info(f"Informe guardado en: {report_path}")

        return report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m quote_engine.importers.excel_products_importer",
        description="Importa Excel de productos a la base de datos EON",
    )
    parser.add_argument("--file", required=True, help="Ruta al archivo Excel")
    parser.add_argument("--force", action="store_true", help="Reimportar aunque el hash ya exista")
    parser.add_argument("--dry-run", action="store_true", dest="dry_run",
                        help="No escribir en DB, solo generar informe")
    parser.add_argument("--report-dir", default="data/reports", dest="report_dir",
                        help="Directorio para el informe (default: data/reports)")
    parser.add_argument("--verbose", "-v", action="store_true", help="Log detallado")

    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    session = None
    if not args.dry_run:
        try:
            from quote_engine.db.session import get_session_factory
            factory = get_session_factory()
            if factory:
                session = factory()
            else:
                logger.warning("DB no configurada. Ejecutando en modo dry-run.")
        except Exception as exc:
            logger.warning(f"No se pudo conectar a DB: {exc}. Modo dry-run.")

    importer = ExcelProductsImporter(db_session=session)

    try:
        report = importer.run(
            file_path=args.file,
            force=args.force,
            report_output_dir=args.report_dir,
        )
    except FileNotFoundError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"Error inesperado: {exc}", file=sys.stderr)
        return 1
    finally:
        if session:
            session.close()

    # Resumen en consola
    print(f"\nImportación completada:")
    print(f"  Filas procesadas:  {report.total_rows}")
    print(f"  Proveedores:       {len(report.suppliers_detected)}")
    print(f"  Productos nuevos:  {report.created_products}")
    print(f"  Precios creados:   {report.created_prices}")
    print(f"  Warnings:          {report.warnings_count}")
    print(f"  Errores:           {report.errors_count}")

    if report.errors_count > 0:
        print("\nErrores:")
        for e in report.errors:
            print(f"  ❌ {e}")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
